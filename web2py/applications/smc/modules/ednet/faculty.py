# -*- coding: utf-8 -*-

from gluon import *
from gluon import current

#import xlrd
# xlrd - quit supporting xlxs files
import openpyxl
import time

from ednet.util import Util
from ednet.appsettings import AppSettings
from ednet.w2py import W2Py
from ednet.ad import AD
from ednet.canvas import Canvas


class Faculty:
    def __init__(self):
        pass
    
    @staticmethod
    def ProcessExcelFile(excel_file):
        out = DIV(Faculty.ClearImportQueue())
        db = current.db # Grab the current db object
        
        #wbook = xlrd.open_workbook(excel_file)
        wbook = openpyxl.load_workbook(excel_file)
        for sheet in wbook:  # wbook.sheets():
            out.append(DIV("Processing Sheet: " + sheet.title))
            out.append(DIV("Processing Enabled Faculty."))
            faculty_account_enabled = True  # status 1 = active, 0 = inactive
            #for row in range(sheet.nrows):
            for row in sheet.iter_rows():
                if Util.GetCellValue(sheet, row, 0).upper() == "USER ID":
                    out.append(DIV("Skipping Header Row."))
                    continue # should be header, skip this line
                if Util.GetCellValue(sheet, row, 0).upper() == "DOC":
                    out.append(DIV("Skipping Header Row."))
                    continue # should be header, skip this line
                if Util.GetCellValue(sheet, row, 0).startswith("**"):
                    # Found stars, switch state
                    faculty_account_enabled = False
                    out.append(DIV("Processing Disabled Faculty."))
                    continue # divider row, skip this line
                if Util.GetCellValue(sheet, row, 0) == "":
                    # out.append(DIV("Skipping Blank Row."))
                    continue # No data, skip this line
                # Expect usernames not numbers
                # if (Util.GetCellValue(sheet, row, 0).isdigit() != True):
                #    out.append(DIV("Skipping non numeric DOC Number: " + Util.GetCellValue(sheet, row, 0)))
                #    continue # No data, skip this line

                # Grab info from cells
                user_id = Util.GetCellValue(sheet, row, 0)
                faculty_name = Util.GetCellValue(sheet, row, 1)
                faculty_password = Util.GetCellValue(sheet, row, 2)
                import_classes = Util.GetCellValue(sheet, row, 3)
                program = Util.GetCellValue(sheet, row, 4)
                additional_fields = ""
                if len(row) > 5:
                    additional_fields = Util.GetJSONFromCellRange(sheet, row, 5, len(row))

                sheet_name = sheet.title
                faculty_guid = None
                account_enabled = faculty_account_enabled
                account_added_on = time.strftime("%c")
                account_updated_on = time.strftime("%c")

                db.faculty_import_queue.insert(
                        user_id=user_id,
                        faculty_name=faculty_name,
                        faculty_password=faculty_password,
                        import_classes=import_classes,
                        program=program,
                        additional_fields=additional_fields,
                        sheet_name=sheet_name,
                        faculty_guid=faculty_guid,
                        account_enabled=account_enabled,
                        account_added_on=account_added_on,
                        account_updated_on=account_updated_on
                    )
                db.commit()
                out.append(DIV("Found: " + user_id))

                # values = []
                # for col in range(sheet.ncols):
                #    values.append(str(sheet.cell(row, col).value))
                # out += ','.join(values)
        out.append(DIV("Processing Complete", _style='font-weight: bold; font-size: 18px; color: red;'))
        return out

    @staticmethod
    def ClearImportQueue():
        db = current.db # Grab the current db object
        # Make sure that the import table is empty
        db.executesql('DELETE FROM faculty_import_queue;')
        return "Faculty Import Queue Cleared."

    @staticmethod
    def CreateW2PyAccounts(sheet_name, override_password=False, override_current_quota=False):
        # Create web2py accounts for this faculty
        db = current.db  # Grab the current db object
        auth = current.auth  # Grab the current auth object
        count = 0
    
        # Get the users to import
        rows = db(db.faculty_import_queue.sheet_name == sheet_name).select()
        for row in rows:
            count += 1
            
            # Gather info for this user
            user_name = Faculty.GetUsername(row.user_id, row=row)
            password = Faculty.GetPassword(row.user_id, row.faculty_password, override_password, row=row)
            user_email = Faculty.GetEmail(row.user_id, row=row)
            (first_name, last_name) = Util.ParseName(row.faculty_name)
            user_ad_quota = Faculty.GetADQuota(row.user_id, override_current_quota)
            user_canvas_quota = Faculty.GetCanvasQuota(row.user_id, override_current_quota)
            
            W2Py.CreateW2PFacultyUser(user_name, password, user_email, first_name, last_name,
                                      user_ad_quota, user_canvas_quota, row)
        return count

    @staticmethod
    def GetLastADLoginTime(faculty_id):
        ret = ""
        db = current.db  # Grab the current db object
        auth = current.auth  # Grab the current auth object
        
        user_name = Faculty.GetUsername(faculty_id)
        
        # AD - Disable
        user_dn = Faculty.GetAD_DN(user_name, Faculty.GetProgram(faculty_id))
        
        ret = AD.GetLastLoginTime(user_dn)
        return ret
    
    @staticmethod
    def DisableAccount(faculty_id):
        ret = ""
        db = current.db  # Grab the current db object
        auth = current.auth  # Grab the current auth object
        
        user_name = Faculty.GetUsername(faculty_id)
        
        pw = "ka&wsa1#74" + str(time.time())
        
        Faculty.SetPassword(faculty_id, pw, False)
        
        # AD - Disable
        user_dn = Faculty.GetAD_DN(user_name, Faculty.GetProgram(faculty_id))
        AD.DisableUser(user_dn)
        
        # Update the database to reflect the change
        db(db.faculty_info.user_id == faculty_id).update(account_enabled=False)
        
        return ret
    
    @staticmethod
    def EnableAccount(faculty_id):
        ret = ""
        db = current.db # Grab the current db object
        auth = current.auth # Grab the current auth object
        
        user_name = Faculty.GetUsername(faculty_id)
        
        pw = Faculty.GetPassword(faculty_id)
        
        Faculty.SetPassword(faculty_id, pw)
        
        # AD - Disable
        user_dn = Faculty.GetAD_DN(user_name, Faculty.GetProgram(faculty_id))
        AD.EnableUser(user_dn)
        
        # Update the database to reflect the change
        db(db.faculty_info.user_id == faculty_id).update(account_enabled=True)
        
        return ret
    
    @staticmethod
    def GetUserIDFromUsername(user_name):
        ret = ""
        db = current.db
        db = current.db  # Grab the current db object
        auth = current.auth  # Grab the current auth object
        
        rows = db(db.auth_user.username == user_name).select()
        for row in rows:
            faculty = db(db.faculty_info.account_id == row.id).select()
            for f in faculty:
                ret = f.user_id
        return ret

    @staticmethod
    def process_config_params(faculty_id, txt, is_username=False, row=None):
        db = current.db  # Grab the current db object
        ret = ""

        # Patters = first_name, last_name, first_name_first_letter, first_name_last_letter,
        # last_name_first_letter, last_name_last_letter
        first_name = ""
        last_name = ""

        first_name_first_letter = ""
        first_name_last_letter = ""
        last_name_first_letter = ""
        last_name_last_letter = ""

        user = db(db.faculty_info.user_id == faculty_id).select(db.faculty_info.faculty_name).first()
        if user is not None:
            (first_name, last_name) = Util.ParseName(user.faculty_name)

        if row is not None:
            (first_name, last_name) = Util.ParseName(row.faculty_name)

        if first_name != "":
            first_name_first_letter = first_name[:1]
            first_name_last_letter = first_name[-1:]
        if last_name != "":
            last_name_first_letter = last_name[:1]
            last_name_last_letter = last_name[-1:]

        ret = txt.replace('<user_id>', faculty_id)
        ret = ret.replace('<first_name>', first_name)
        ret = ret.replace('<first_name_first_letter>', first_name_first_letter)
        ret = ret.replace('<first_name_last_letter>', first_name_last_letter)
        ret = ret.replace('<last_name>', last_name)
        ret = ret.replace('<last_name_first_letter>', last_name_first_letter)
        ret = ret.replace('<last_name_last_letter>', last_name_last_letter)

        # Make sure we don't call this if it is coming FROM the username function
        if is_username is not True:
            user_name = Faculty.GetUsername(faculty_id)
            ret = ret.replace('<user_name>', user_name)
            ret = ret.replace('%username%', user_name)

        return str(ret)

    @staticmethod
    def GetUsername(faculty_id, row=None):
        db = current.db
        ret = ""

        # Try to get the real (current) username for this id
        row = db(db.faculty_info.user_id.like(faculty_id)).select().first()
        if row:
            # Got the faculty record, get the matching user id
            user_name = row.account_id.username
            if user_name is not None and user_name != "":
                #print("Found real user name: " + user_name)
                ret = user_name
        # Unable to find that, return the default pattern
        if ret == "":
            pattern = AppSettings.GetValue('faculty_id_pattern', '<user_id>')
            ret = Faculty.process_config_params(faculty_id, pattern, is_username=True, row=row)
        
        return str(ret)
    
    @staticmethod
    def GetProgram(faculty_id):
        db = current.db
        ret = ""
        user = db(db.faculty_info.user_id==faculty_id).select(db.faculty_info.program).first()
        if user is not None:
            ret = user.program
        if ret is None:
            ret = ""
        return ret

    @staticmethod
    def GetPassword(faculty_id, import_password="", override_password=False, row=None):
        db = current.db
        # Get current password if faculty exists so we don't
        # reset it if the faculty has already changed it unless override_password == True
        ret = import_password
        if override_password is not True:
            rows = db(db.faculty_info.user_id == faculty_id).select(db.faculty_info.faculty_password)
            for row2 in rows:
                ret = row2.faculty_password
        # If password isn't set, use the pattern and set the default pw
        # Default - add SID to beginning and ! to end (e.g. SID1888182!)
        if ret == "":
            pattern = AppSettings.GetValue('faculty_password_pattern', 'FID<user_id>!')
            ret = Faculty.process_config_params(faculty_id, pattern, row=row)
        return str(ret)
    
    @staticmethod
    def SetPassword(faculty_id, new_password, update_db=True):
        db = current.db
        ret = ""
        
        if faculty_id == "" or new_password == "":
            return "Can't set empty password!"
        
        if len(new_password) < 6:
            return "Can't set password - Too short!"
        
        # Set AD password
        user_name = Faculty.GetUsername(faculty_id)
        if AD.Connect() is True:
            user_dn = Faculty.GetAD_DN(user_name, Faculty.GetProgram(faculty_id))
            if AD.SetPassword(user_dn, new_password) is not True:
                return "Error setting AD password: " + AD.GetErrorString()
        else:
            # Don't set pw if AD is disabled
            print("skipping ad set")
            pass
        
        # Set Canvas password
        if Canvas.SetPassword(user_name, new_password) is not True:
            return "Error setting Canvas password: " + Canvas.GetErrorString()
        
        if W2Py.SetFacultyPassword(user_name, new_password, update_db) is not True:
            return "Error setting SMC password"
        
        return ret

    @staticmethod
    def GetEmail(faculty_id, row=None):
        # Default - append username and @domain.com (e.g. bobsmith@correctionsed.com)
        pattern = AppSettings.GetValue('faculty_email_pattern', '<user_name>@correctionsed.com')
        ret = Faculty.process_config_params(faculty_id, pattern, row=row)
        return str(ret)
    
    @staticmethod
    def GetHomeDirectory(faculty_id):
        # Default - is empty  # \\files\faculty\%username%
        pattern = AppSettings.GetValue('ad_faculty_home_directory', '')
        ret = Faculty.process_config_params(faculty_id, pattern)
        return str(ret)

    @staticmethod
    def GetHomeDrive(faculty_id):
        # Default - nothing (off)
        pattern = AppSettings.GetValue('ad_faculty_home_drive', '')
        ret = Faculty.process_config_params(faculty_id, pattern)
        return str(ret)

    @staticmethod
    def GetLoginScriptPath(faculty_id):
        # Default - nothing (off)
        pattern = AppSettings.GetValue('ad_faculty_login_script_path', '')
        ret = Faculty.process_config_params(faculty_id, pattern)
        return str(ret)

    @staticmethod
    def GetProfilePath(faculty_id):
        # Default - nothing (off)
        pattern = AppSettings.GetValue('ad_faculty_profile_directory', '')
        ret = Faculty.process_config_params(faculty_id, pattern)
        return str(ret)

    @staticmethod
    def GetCanvasQuota(faculty_id, override_quota=False):
        # Lookup the quota for this user or use the default
        db = current.db

        # Start with a default value
        quota = AppSettings.GetValue('canvas_faculty_quota', '1048576')
        # Query for the faculty value
        if override_quota is not True:
            rows = db(db.faculty_info.user_id == faculty_id).select(db.faculty_info.faculty_canvas_quota)
            for row in rows:
                quota = row.faculty_canvas_quota
        
        return str(quota)
    
    @staticmethod
    def GetADQuota(faculty_id, override_quota=False):
        # Lookup the quota for this user or use the default
        db = current.db

        # Start with a default value
        quota = AppSettings.GetValue('ad_faculty_home_directory_quota', '1048576')
        # Query for the faculty value
        if override_quota is not True:
            rows = db(db.faculty_info.user_id == faculty_id).select(db.faculty_info.faculty_ad_quota)
            for row in rows:
                quota = row.faculty_ad_quota
        
        return str(quota)

    @staticmethod
    def QueueActiveDirectoryImports(sheet_name):
        db = current.db  # Grab the current db object
        count = 0  # The number of entries processed

        ldap_enabled = AppSettings.GetValue('ad_import_enabled', False)
        if ldap_enabled is not True:
            return count

        # Clear the queue and add an entry in the queue table for
        # each faculty
        db.executesql('DELETE FROM faculty_ad_import_queue;')
        db.executesql('DELETE FROM faculty_ad_import_status;')

        rows = db(db.faculty_import_queue.sheet_name == sheet_name).select(db.faculty_import_queue.id)
        for row in rows:
            count += 1
            db.faculty_ad_import_queue.insert(faculty_import_queue=row['id'])

        return count

    @staticmethod
    def QueueCanvasImports(sheet_name):
        db = current.db # Grab the current db object
        count = 0 # The number of entries processed

        canvas_enabled = AppSettings.GetValue('canvas_import_enabled', False)
        if canvas_enabled is not True:
            return count

        # Clear the queue and add an entry in the queue table for
        # each faculty
        db.executesql('DELETE FROM faculty_canvas_import_queue;')
        db.executesql('DELETE FROM faculty_canvas_import_status;')

        rows = db(db.faculty_import_queue.sheet_name==sheet_name).select(db.faculty_import_queue.id)
        for row in rows:
            count += 1
            db.faculty_canvas_import_queue.insert(faculty_import_queue=row['id'])

        return count

    @staticmethod
    def GetEnrolledClassesString(user_id):
        db = current.db # Grab the current db object
        ret = ""
        
        classes = Faculty.GetEnrolledClasses(user_id)
        for c in classes:
            if ret != "":
                ret += ", "
            ret += c.course_code

        if ret == "":
            ret = "No Classes"
        return ret
    
    @staticmethod
    def GetEnrolledClasses(user_id):
        db = current.db # Grab the current db object
        ret = []
        
        user = db(db.faculty_info.user_id==user_id).select().first()
        if user is not None:
            classes = user.faculty_enrollment.select()
            for c in classes:
                if c.enrollment_status == "active":
                    ret.append(c)
        return ret
    
    @staticmethod
    def GetAD_DN(user_name, program):
        ret = Faculty.GetAD_CN(program)
        ret = AD.GetDN(user_name, ret)
        
        return ret
    
    @staticmethod
    def GetAD_CN(program):
        ad_cn = AppSettings.GetValue('ad_faculty_cn', 'OU=Faculty,DC=ad,DC=correctionsed,DC=com')
        if program == "" or program is None:
            # If no program, take off the program OU
            ret = ad_cn.replace('ou=', 'OU=')
            ret = ret.replace('OU=<program>,', '')
        else:
            ret = ad_cn.replace('<program>', program)
        
        return ret

    @staticmethod
    def AddClass(user_id, class_name):
        db = current.db
        user = db(db.faculty_info.user_id == user_id).select().first()
        if user is not None:
            uid = user['id']
            if db((db.faculty_enrollment.parent_id == uid) &
                  (db.faculty_enrollment.course_code == class_name)).select().first() is None:
                db.faculty_enrollment.insert(parent_id=uid, course_code=class_name, enrolled_on=time.strftime("%c"),
                                             enrollment_status='active')
    
    @staticmethod
    def ProcessADFaculty():
        db = current.db # Grab the current db object
        scheduler = current.scheduler
        ret = ""

        ldap_enabled = AppSettings.GetValue('ad_import_enabled', False)
        if ldap_enabled is not True:
            return "Done! - LDAP Import Disabled"
        
        if AD.Connect() is not True:
            ret += "<b>Error connecting to Active Directory server</b><br/><font size=-4>"
            ret += AD.GetErrorString()
            ret += "</font><br/>Done!"
            return ret
        
        if AD.VerifyADSettings() is not True:
            ret += "<b>Error verifying AD settings</b><br/><font size=-4>"
            ret += AD.GetErrorString()
            ret += "</font><br/>Done!"
            return ret
        else:
            # If everything is good clear errors
            AD._errors = []

        ad_faculty_group_cn = AppSettings.GetValue('ad_faculty_group_cn',
                                                   'OU=FacultyGroups,DC=ad,DC=correctionsed,DC=com')
        ad_faculty_group_dn = 'CN=Faculty,' + ad_faculty_group_cn
        
        # Ensure the faculty group exists
        if AD.CreateGroup(ad_faculty_group_dn) is not True:
            ret += "<b>Error creating faculty group:</b> " + str(ad_faculty_group_dn) + "<br />"
            ret += str(AD._errors)
        
        # Grab the first faculty off the queue
        rows = db(db.faculty_import_queue.id == db.faculty_ad_import_queue.faculty_import_queue)\
            .select(orderby=db.faculty_import_queue.account_enabled|db.faculty_import_queue.faculty_name,
                    limitby=(0, 1))
        
        for row in rows:
            # Pop the faculty off the queue
            db(db.faculty_ad_import_queue.id == row.faculty_ad_import_queue.id).delete()
            db.commit()
            # Get the faculty info
            faculty_user_name = Faculty.GetUsername(row.faculty_import_queue.user_id)
            faculty_password = Faculty.GetPassword(row.faculty_import_queue.user_id, row.faculty_import_queue.faculty_password)
            (faculty_first_name, faculty_last_name) = Util.ParseName(row.faculty_import_queue.faculty_name)
            faculty_email = Faculty.GetEmail(row.faculty_import_queue.user_id)
            faculty_display_name = row.faculty_import_queue.faculty_name + " (" + faculty_user_name + ")"
            faculty_user_id = row.faculty_import_queue.user_id
            faculty_home_directory = Faculty.GetHomeDirectory(row.faculty_import_queue.user_id)
            faculty_home_drive = Faculty.GetHomeDrive(row.faculty_import_queue.user_id)
            faculty_login_script_path = Faculty.GetLoginScriptPath(row.faculty_import_queue.user_id)
            faculty_profile_path = Faculty.GetProfilePath(row.faculty_import_queue.user_id)
            faculty_enabled = row.faculty_import_queue.account_enabled
            faculty_quota = Faculty.GetADQuota(row.faculty_import_queue.user_id)
            faculty_dn = Faculty.GetAD_DN(faculty_user_name, row.faculty_import_queue.program)
            faculty_cn = Faculty.GetAD_CN(row.faculty_import_queue.program)
            
            first_run = False
            fr = db(db.faculty_ad_import_status.user_id==row.faculty_import_queue.user_id).select().first()
            if fr is None:
                first_run = True
            db.faculty_ad_import_status.insert(user_id=row.faculty_import_queue.user_id)
            db.commit()
            
            # Create the faculty
            if AD.CreateUser(faculty_user_name, faculty_cn) is not True:
                ret += "<b>Error creating faculty account:</b> " + str(faculty_user_name) + " - " + str(faculty_cn) +\
                       "<br />Done!"  # + AD.GetErrorString()
                return ret
            db.commit()
            # Update user with current info
            if AD.UpdateUserInfo(faculty_dn, email_address=faculty_email, first_name=faculty_first_name,
                                 last_name=faculty_last_name, display_name=faculty_display_name,
                                 description="Faculty Account", id_number=faculty_user_name,
                                 home_drive_letter=faculty_home_drive, home_directory=faculty_home_directory,
                                 login_script=faculty_login_script_path, profile_path=faculty_profile_path,
                                 ts_allow_login='TRUE') is not True:
                ret += "<b>Error creating setting faculty information:</b> " + str(faculty_user_name) + "<br />"
            db.commit()
            # Set password
            if AD.SetPassword(faculty_dn, faculty_password) is not True:
                ret += "<b>Error setting password for user:</b> " + str(faculty_user_name) + "<br />"
            db.commit()
            # Add to the faculty group
            if AD.AddUserToGroup(faculty_dn, ad_faculty_group_dn) is not True:
                ret += "<b>Error adding user to faculty group:</b> " + str(faculty_user_name) + "<br />"
            db.commit()
            if faculty_enabled is True:
                AD.EnableUser(faculty_dn)
            else:
                AD.DisableUser(faculty_dn)
            db.commit()

            # NOTE - We do NOT want to remove faculty from anything automatically
            
            # Get the list of classes for this faculty
            if faculty_enabled is True:
                enroll_classes = row.faculty_import_queue.import_classes.split(',')
                for enroll_class in enroll_classes:
                    # Trim spaces
                    enroll_class = enroll_class.strip()

                    if enroll_class == '':
                        continue  # Skip empty class names
                    # ret += "Enrolling into: " + enroll_class
                    Faculty.AddClass(row.faculty_import_queue.user_id, enroll_class)
                    
                    class_dn = AD.GetDN(enroll_class + "-F", ad_faculty_group_cn)
                    if AD.GetLDAPObject(class_dn) is None:
                        # Class group doesn't exist, add it
                        if AD.CreateGroup(class_dn) is not True:
                            ret += "<b>Error creating class group:</b> " + str(enroll_class) + "<br />"

                    # Add faculty to the class group
                    if AD.AddUserToGroup(faculty_dn, class_dn) is not True:
                        ret += "<b>Error adding faculty to group:</b> " + str(faculty_user_name) +\
                               "/" + str(enroll_class) + "<br />"
            db.commit()
            # Setup physical home directory
            if faculty_enabled is True:
                # if (AD.CreateHomeDirectory(faculty_user_name, faculty_home_directory) != True):
                #    ret += "<b>Error creating home folder:</b> " + str(faculty_user_name) + "<br />"
                if first_run:
                    result = scheduler.queue_task('create_home_directory',
                                                  pvars=dict(user_name=faculty_user_name,
                                                             home_directory=faculty_home_directory),
                                                  timeout=1200, immediate=True, sync_output=5,
                                                  group_name="create_home_directory")
                if AD.SetDriveQuota(faculty_user_name, faculty_quota) is not True:
                    ret += "<b>Error setting quota for faculty:</b> " + str(faculty_user_name) + "<br />"
            db.commit()
            # Show errors
            if len(AD._errors) > 0:
                ret += AD.GetErrorString()
            
            ret += faculty_display_name  # + " (" + faculty_user_name + ")"
            if row.faculty_import_queue.account_enabled is True:
                ret += " - <span style='color: green; font-weight: bolder;'>Imported</span>"
            else:
                ret += " - <span style='color: red; font-weight: bolder;'>Disabled</span>"

        # Finished importing, clean up after AD
        AD.Close()
        
        if ret == "":
            ret = "Done!"
        return ret
    
    @staticmethod
    def ProcessCanvasFaculty():
        db = current.db  # Grab the current db object
        ret = ""
        log = ""
        Canvas.Close()
        
        if Canvas.Connect() is not True:
            ret += "<b>Error connecting to Canvas server</b><br/><font size=-4>"
            ret += Canvas.GetErrorString()
            ret += "</font><br/>"
            ret += "Done!"
            return ret
        
        if Canvas._canvas_enabled is not True or Canvas._canvas_import_enabled is not True:
            return "Done! - Canvas Import Disabled"

        # Pop one off the queue
        rows = db(db.faculty_import_queue.id == db.faculty_canvas_import_queue.faculty_import_queue) \
            .select(orderby=db.faculty_import_queue.account_enabled | db.faculty_import_queue.faculty_name,
                    limitby=(0, 1))
        for row in rows:
            # remove item from the queue
            db(db.faculty_canvas_import_queue.id == row.faculty_canvas_import_queue.id).delete()
            db.commit()
            # Get info for current faculty
            faculty_user_name = Faculty.GetUsername(row.faculty_import_queue.user_id)
            faculty_password = Faculty.GetPassword(row.faculty_import_queue.user_id,
                                                   row.faculty_import_queue.faculty_password)
            (faculty_first_name, faculty_last_name) = Util.ParseName(row.faculty_import_queue.faculty_name)
            faculty_email = Faculty.GetEmail(row.faculty_import_queue.user_id)
            faculty_display_name = row.faculty_import_queue.faculty_name
            faculty_user_id = row.faculty_import_queue.user_id
            faculty_enabled = row.faculty_import_queue.account_enabled
            db.commit()
            
            first_run = False
            fr = db(db.faculty_canvas_import_status.user_id == row.faculty_import_queue.user_id).select().first()
            if fr is None:
                first_run = True
            db.faculty_canvas_import_status.insert(user_id=row.faculty_import_queue.user_id)
            db.commit()
            
            # Make sure the user exists
            canvas_user = Canvas.CreateUser(faculty_user_name, faculty_password, faculty_first_name,
                                            faculty_last_name, faculty_email)
            if canvas_user is None:
                ret += "<b>Error creating account:</b> " + str(faculty_user_name) +\
                       "<br />" + Canvas.GetErrorString() + "<br />Done!"
                return ret
            
            # Mark all current classes as complete and delete teacher enrollment
            # NOTE - Disabled so that faculty aren't removed from their classes
            # if (first_run == True):
            #    Canvas.CompleteAllClasses(faculty_user_name, True)
            # ret += "<b>Current Enrollment: </b>" + str(Canvas.GetCurrentClasses(faculty_user_name))
            
            # The list of classes from the import
            enroll_classes = row.faculty_import_queue.import_classes.split(',')
            class_str = ""            
            if faculty_enabled is True:
                for enroll_class in enroll_classes:
                    if class_str != "":
                        class_str += ", "
                    class_str += enroll_class
                    Canvas.EnrollTeacher(canvas_user, enroll_class)
                    Faculty.AddClass(row.faculty_import_queue.user_id, enroll_class)
                    db.commit()
            
            ret += faculty_display_name + " (" + faculty_user_name + ")"
            if row.faculty_import_queue.account_enabled is True:
                ret += " - <span style='color: green; font-weight: bolder;'>Imported (" + class_str + ")</span>"
            else:
                ret += " - <span style='color: red; font-weight: bolder;'>Disabled</span>"

        # Show errors
        if len(Canvas._errors) > 0:
            ret += Canvas.GetErrorString()

        if ret == "":
            ret = "Done!"
            
        return ret
